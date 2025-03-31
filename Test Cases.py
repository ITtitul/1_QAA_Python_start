import pandas as pd

# Создаем DataFrame с колонками для тест-кейсов
test_cases = pd.DataFrame(columns=[
    'ID',
    'Название теста',
    'Приоритет',
    'Предусловия',
    'Шаги',
    'Ожидаемый результат',
    'Фактический результат',
    'Статус',
    'Окружение',
    'Автоматизирован'
])

# Добавляем пример тест-кейса
test_cases.loc[0] = [
    'TC_001',
    'Успешная авторизация с валидными данными',
    'Высокий',
    '1. Открыта страница авторизации\n2. Система доступна',
    '1. Ввести логин "admin"\n2. Ввести пароль "admin_password"\n3. Нажать кнопку "Войти"',
    'Отображается сообщение "Успешный вход!"',
    '',
    'Не выполнен',
    'Chrome 122.0',
    'Да'
]

# Сохраняем в Excel
test_cases.to_excel('test_cases.xlsx', index=False)

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font

# Создаем и сохраняем базовый файл
test_cases.to_excel('test_cases.xlsx', index=False)

# Открываем для форматирования
wb = load_workbook('test_cases.xlsx')
ws = wb.active

# Форматирование заголовков
header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
header_font = Font(color='FFFFFF', bold=True)

# Применяем стили к заголовкам
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font

# Выравнивание и перенос текста
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical='top')

# Устанавливаем ширину колонок
column_widths = {
    'A': 10,  # ID
    'B': 30,  # Название теста
    'C': 15,  # Приоритет
    'D': 30,  # Предусловия
    'E': 40,  # Шаги
    'F': 30,  # Ожидаемый результат
    'G': 30,  # Фактический результат
    'H': 15,  # Статус
    'I': 20,  # Окружение
    'J': 15   # Автоматизирован
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# Сохраняем отформатированный файл
wb.save('test_cases.xlsx')